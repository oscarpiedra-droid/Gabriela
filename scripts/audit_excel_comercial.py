"""
audit_excel_comercial.py
========================
Auditoría COMPLETA del sistema de validación comercial BUR2000.

EXIT CODE:
  0  → Todo correcto (0 errores)
  1  → Hay errores críticos

Uso:
  miniconda\python.exe -X utf8 scripts\audit_excel_comercial.py
"""

import sys
import os
import json
import re
import math

# ── Setup de paths ───────────────────────────────────────────────────────────
_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(_ROOT, "app"))

import pandas as pd

# ── Constantes del Excel ─────────────────────────────────────────────────────
EXCEL         = os.path.join(_ROOT, "Nuevo", "ENERO 2026 - Con Axarquia.xlsx")
SHEET_COND    = "Condiciones de dtos Enero 2026"
SHEET_PORTES  = "Portes Abril 2026"
HEADER_ROW    = 16   # 0-based (fila 17 en Excel)

# Segmentos canónicos — deben coincidir exactamente con el Excel
SEGMENTOS_CANONICOS = [
    "Almacenes Especialistas (PYL)",
    "Almacenes Generalistas",
    "Empresas Constructoras",
    "Empresas Instaladoras",
    "Almacenes e Instaladores (Gama SOUND)",
    "Axarquía de Aislamientos (Distribución)",
]

# ── Estado global de la auditoría ────────────────────────────────────────────
_errors   = []
_warnings = []
_ok_count = 0

def ok(label, detail=""):
    global _ok_count
    _ok_count += 1
    detail_str = f" [{detail}]" if detail else ""
    print(f"  [OK  ] {label}{detail_str}")

def err(label, detail=""):
    _errors.append(f"{label}: {detail}")
    print(f"  [ERR ] {label} — {detail}")

def warn(label, detail=""):
    _warnings.append(f"{label}: {detail}")
    print(f"  [WARN] {label}" + (f" — {detail}" if detail else ""))


# ─────────────────────────────────────────────────────────────────────────────
# BLOQUE 1 — Integridad del archivo Excel
# ─────────────────────────────────────────────────────────────────────────────
def bloque_1_integridad():
    print("\n[1] INTEGRIDAD DEL ARCHIVO EXCEL")
    if not os.path.exists(EXCEL):
        err("Archivo Excel existe", EXCEL)
        return False
    ok("Archivo Excel existe", os.path.basename(EXCEL))

    try:
        from openpyxl import load_workbook
        wb = load_workbook(EXCEL, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        err("Abrir Excel con openpyxl", str(e))
        return False

    if SHEET_COND in sheets:
        ok(f"Hoja '{SHEET_COND}' presente")
    else:
        err(f"Hoja '{SHEET_COND}' AUSENTE", f"Hojas disponibles: {sheets}")

    if SHEET_PORTES in sheets:
        ok(f"Hoja '{SHEET_PORTES}' presente")
    else:
        err(f"Hoja '{SHEET_PORTES}' AUSENTE", f"Hojas disponibles: {sheets}")

    return True


# ─────────────────────────────────────────────────────────────────────────────
# BLOQUE 2 — Lectura de condiciones
# ─────────────────────────────────────────────────────────────────────────────
def bloque_2_lectura():
    print("\n[2] LECTURA HOJA CONDICIONES (fila 17, cols D:I)")
    try:
        df = pd.read_excel(EXCEL, sheet_name=SHEET_COND, header=HEADER_ROW,
                           usecols="D:I", engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how="all")
        df = df[df["Segmento"].notna() & df["Familia"].notna()]
        ok("DataFrame cargado", f"{len(df)} filas de datos")
    except Exception as e:
        err("Leer hoja condiciones", str(e))
        return None

    # Columnas esperadas
    cols_esperadas = [
        "Segmento", "Familia", "Tramo facturación",
        "DTO Territorial (%)", "DTO Baleares (%)",
        "Condición mínima (familias/referencias)",
    ]
    all_present = True
    for col in cols_esperadas:
        if col in df.columns:
            ok(f"Columna '{col}'")
        else:
            err(f"Columna '{col}' AUSENTE", f"Columnas detectadas: {list(df.columns)}")
            all_present = False

    if not all_present:
        return None

    return df


# ─────────────────────────────────────────────────────────────────────────────
# BLOQUE 3 — Segmentos
# ─────────────────────────────────────────────────────────────────────────────
def bloque_3_segmentos(df):
    print("\n[3] SEGMENTOS EN EXCEL")
    segs_excel = list(df["Segmento"].dropna().unique())
    print(f"  Segmentos encontrados ({len(segs_excel)}):")
    for s in sorted(segs_excel):
        print(f"    - '{s}'")

    # Verificar que cada canónico está presente (comparación strip exacta)
    for seg in SEGMENTOS_CANONICOS:
        encontrado = any(s.strip() == seg.strip() for s in segs_excel)
        if encontrado:
            ok(f"Segmento canonico '{seg[:50]}'")
        else:
            err(f"Segmento canonico AUSENTE", f"'{seg}' no encontrado en Excel")

    # Avisar si hay segmentos extra no reconocidos
    for s in segs_excel:
        if s.strip() not in [c.strip() for c in SEGMENTOS_CANONICOS]:
            warn("Segmento extra no reconocido", f"'{s}' — añadir al catalogo si es activo")

    return set(s.strip() for s in segs_excel)


# ─────────────────────────────────────────────────────────────────────────────
# BLOQUE 4 — Tramos por segmento
# ─────────────────────────────────────────────────────────────────────────────
def bloque_4_tramos(df):
    print("\n[4] TRAMOS DE FACTURACION POR SEGMENTO")

    def parse_tramo_min(v):
        try:
            return float(v)
        except Exception:
            return 0.0

    n_errores_mono = 0
    for seg, grp in df.groupby("Segmento"):
        tramos   = grp["Tramo facturación"].unique()
        familias = grp["Familia"].nunique()
        n_tramos = len(tramos)
        axarquia = "xarqu" in seg.lower()
        print(f"  {seg[:55]}: {n_tramos} tramos, {familias} familias")

        if axarquia and n_tramos < 8:
            warn(f"Axarquia deberia tener 8 tramos", f"tiene {n_tramos} en segmento '{seg}'")

        # Monotonía DTO por (segmento, familia)
        for fam, grp2 in grp.groupby("Familia"):
            grp2 = grp2.copy()
            grp2["_tmin"] = grp2["Tramo facturación"].apply(parse_tramo_min)
            grp2 = grp2.sort_values("_tmin")
            dtos = grp2["DTO Territorial (%)"].dropna().values
            for i in range(1, len(dtos)):
                if dtos[i] < dtos[i-1] - 0.01:
                    warn("Monotonia rota (mayor tramo no implica mayor DTO)",
                         f"{seg[:30]}/{fam}: {dtos[i-1]}% -> {dtos[i]}%")
                    n_errores_mono += 1

    if n_errores_mono == 0:
        ok("Monotonia DTO correcta en todos los segmentos/familias")


# ─────────────────────────────────────────────────────────────────────────────
# BLOQUE 5 — Valores DTO
# ─────────────────────────────────────────────────────────────────────────────
def bloque_5_dtos(df):
    print("\n[5] VALORES DTO — INTEGRIDAD NUMERICA")

    dto_ter = "DTO Territorial (%)"
    dto_bal = "DTO Baleares (%)"
    axarquia_mask = df["Segmento"].str.contains("xarqu", case=False, na=False)

    # DTO Territorial nunca nulo
    nan_ter = df[df[dto_ter].isna()]
    if len(nan_ter) == 0:
        ok(f"DTO Territorial sin NaN ({len(df)} filas)")
    else:
        err(f"DTO Territorial con NaN", f"{len(nan_ter)} filas: {nan_ter[['Segmento','Familia']].head(3).to_dict('records')}")

    # DTO Baleares: NaN solo en Axarquía
    nan_bal = df[df[dto_bal].isna()]
    nan_bal_non_axq = nan_bal[~axarquia_mask.loc[nan_bal.index]]
    if len(nan_bal_non_axq) == 0:
        ok("DTO Baleares NaN solo en Axarquia")
    else:
        err("DTO Baleares NaN en segmentos no-Axarquia",
            str(nan_bal_non_axq[["Segmento", "Familia"]].head(5).to_dict("records")))

    # Rango [0, 100]
    for col in [dto_ter, dto_bal]:
        valid = df[df[col].notna()]
        out   = valid[~valid[col].between(0, 100)]
        if len(out) == 0:
            ok(f"{col} en rango [0, 100]")
        else:
            err(f"{col} fuera de rango",
                f"{len(out)} valores: {out[[col,'Segmento','Familia']].head(3).to_dict('records')}")

    # NaN en Axarquía DTO Baleares es esperado (distribuidor peninsular)
    nan_axq = nan_bal[axarquia_mask.loc[nan_bal.index]]
    if len(nan_axq) > 0:
        ok(f"Axarquia DTO Baleares=NaN (esperado, distribuidor peninsular)", f"{len(nan_axq)} filas")


# ─────────────────────────────────────────────────────────────────────────────
# BLOQUE 6 — Hoja Portes
# ─────────────────────────────────────────────────────────────────────────────
def bloque_6_portes():
    print("\n[6] HOJA PORTES ABRIL 2026")
    try:
        dp = pd.read_excel(EXCEL, sheet_name=SHEET_PORTES, header=9,
                           usecols="B:D", engine="openpyxl")
        dp.columns = [str(c).strip() for c in dp.columns]
        dp = dp.dropna(how="all").dropna(subset=[dp.columns[0]])
        ok("Hoja Portes cargada", f"{len(dp)} gamas")

        cols_esperadas = ["Gama", "Portes Gratis Desde", "Portes por Comunidad Autónoma"]
        for col in cols_esperadas:
            if col in dp.columns:
                ok(f"Columna Portes '{col}'")
            else:
                warn("Columna Portes no encontrada exactamente", f"'{col}' — columnas: {list(dp.columns)}")

        if len(dp) >= 5:
            ok("Portes tiene al menos 5 gamas de producto")
        else:
            warn("Portes tiene pocas gamas", f"{len(dp)} (esperado >= 5)")

    except Exception as e:
        err("Leer hoja Portes", str(e))


# ─────────────────────────────────────────────────────────────────────────────
# BLOQUE 7 — Catálogo JSON de Homologación
# ─────────────────────────────────────────────────────────────────────────────
def bloque_7_catalogo(segmentos_excel):
    print("\n[7] CATALOGO JSON HOMOLOGACION")
    catalog_path = os.path.join(_ROOT, "app", "db", "services", "homologacion_clientes.json")

    if not os.path.exists(catalog_path):
        err("Catalogo JSON existe", catalog_path)
        return

    try:
        with open(catalog_path, encoding="utf-8") as f:
            catalog = json.load(f)
        ok("JSON valido y legible")
    except json.JSONDecodeError as e:
        err("JSON no es valido", str(e))
        return

    version  = catalog.get("_meta", {}).get("version", "?")
    entries  = catalog.get("homologacion", [])
    active   = [e for e in entries if e.get("estado") == "activo"]
    ok(f"Entradas activas", f"{len(active)} (version {version})")

    # Axarquia cubierta
    axq_entries = [e for e in active if "xarqu" in e.get("odoo_tipo_cliente", "").lower()]
    if len(axq_entries) >= 3:
        ok("Axarquia cubierta en catalogo", f"{len(axq_entries)} variantes")
    else:
        err("Axarquia tiene pocas entradas en catalogo",
            f"{len(axq_entries)} (minimo 3: con tilde, sin tilde, alias punto)")

    # Todos los segmentos del Excel tienen al menos 1 entrada
    segs_cubiertos = {e.get("segmento_aplicacion", "").strip() for e in active}
    for seg in segmentos_excel:
        seg_strip = seg.strip()
        if seg_strip in segs_cubiertos:
            ok(f"Segmento '{seg_strip[:50]}' cubierto")
        else:
            err(f"Segmento SIN cobertura en catalogo", f"'{seg_strip}'")

    # Nombres de segmento en catálogo usan capitalización exacta del Excel
    for e in active:
        s = e.get("segmento_aplicacion", "").strip()
        if s and s not in segs_cubiertos and s not in [
            "Condiciones fuera de la tabla", "Tipo de Empresa por Definir", "Especialistas"
        ]:
            warn("segmento_aplicacion en catalogo no reconocido", f"'{s}'")


# ─────────────────────────────────────────────────────────────────────────────
# BLOQUE 8 — Motor de validación (tests funcionales)
# ─────────────────────────────────────────────────────────────────────────────
def bloque_8_motor_validacion():
    print("\n[8] TEST FUNCIONAL — MOTOR DE VALIDACION")
    try:
        from db.services.commercial_conditions_service import (
            DiscountProposalService, COL_SEGMENTO, COL_FAMILIA
        )
        svc = DiscountProposalService(excel_path=EXCEL)
        records = svc.get_proposal_data()
        ok("DiscountProposalService carga datos", f"{len(records)} reglas")
    except Exception as e:
        err("DiscountProposalService importable", str(e))
        return

    # Helper para buscar un (seg, fam) real del Excel
    def find_sf(seg_contains):
        for r in records:
            if seg_contains.lower() in str(r.get(COL_SEGMENTO, "")).lower():
                return r.get(COL_SEGMENTO), r.get(COL_FAMILIA)
        return None, None

    tests = [
        # (descripcion, segmento, familia, importe, territorio, dto, status_esperado)
    ]

    # Test 1 — PYL DTO=0 → OK
    s, f = find_sf("Almacenes Especialistas")
    if s and f:
        r = svc.validate_range(s, f, 5000, "Madrid", 0)
        if r["status"] in ("OK", "AVISO"):
            ok("PYL DTO=0 -> OK/AVISO", r["status"])
        else:
            err("PYL DTO=0 deberia ser OK/AVISO", r["status"])
    else:
        warn("No se encontro segmento PYL para test funcional")

    # Test 2 — PYL DTO=99% → BLOQUEADO
    if s and f:
        r2 = svc.validate_range(s, f, 5000, "Madrid", 99)
        if r2["status"] in ("BLOQUEADO", "AVISO"):
            ok("PYL DTO=99% -> BLOQUEADO/AVISO", r2["status"])
        else:
            err("PYL DTO=99% deberia ser BLOQUEADO", r2["status"])

    # Test 3 — Axarquia + Baleares → fallback a territorial, NO nulo
    s_axq, f_axq = find_sf("xarqu")
    if s_axq and f_axq:
        r3 = svc.validate_range(s_axq, f_axq, 1000, "Baleares", 0)
        dto_max = r3.get("rules", {}).get("max")
        if r3["status"] == "ERROR":
            err("Axarquia+Baleares devuelve ERROR (no deberia)", str(r3))
        elif dto_max is None or (isinstance(dto_max, float) and math.isnan(dto_max)):
            err("Axarquia+Baleares DTO max es NaN/None", f"dto_max={dto_max}")
        else:
            ok("Axarquia+Baleares fallback a territorial", f"status={r3['status']} dto_max={dto_max}")
    else:
        warn("No se encontro segmento Axarquia para test Baleares")

    # Test 4 — Segmento inexistente → OK (no bloquear)
    r4 = svc.validate_range("SEGMENTO_INEXISTENTE_XYZ", "FAMILIA_XYZ", 5000, "Madrid", 50)
    if r4["status"] == "OK":
        ok("Segmento inexistente -> OK (no bloquea)", r4.get("msg", ""))
    else:
        err("Segmento inexistente deberia ser OK", r4["status"])

    # Test 5 — get_portes_data
    portes = svc.get_portes_data()
    if len(portes) > 0:
        ok("get_portes_data devuelve datos", f"{len(portes)} gamas")
    else:
        err("get_portes_data vacio")

    # Test 6 — get_dto_for
    if s and f:
        dto = svc.get_dto_for(s, f, 5000, "Madrid")
        if dto is not None and 0 <= dto <= 100:
            ok("get_dto_for devuelve valor valido", f"{dto}%")
        else:
            warn("get_dto_for devuelve valor inesperado", f"dto={dto}")

    # Test 7 — Invalidar cache y recargar
    svc.invalidate_cache()
    records2 = svc.get_proposal_data()
    if len(records2) == len(records):
        ok("Invalidar cache y recargar", f"{len(records2)} reglas == original")
    else:
        warn("Cache reload devuelve diferente numero de reglas",
             f"original={len(records)} reload={len(records2)}")


# ─────────────────────────────────────────────────────────────────────────────
# BLOQUE 9 — Motor de homologación (tests funcionales)
# ─────────────────────────────────────────────────────────────────────────────
def bloque_9_homologacion():
    print("\n[9] TEST FUNCIONAL — MOTOR DE HOMOLOGACION")
    try:
        from db.services.homologacion_service import HomologacionService, HomologacionStatus
        homo = HomologacionService()
        ok("HomologacionService importado correctamente")
    except Exception as e:
        err("HomologacionService importable", str(e))
        return

    # Tipos reales de Odoo (sacados de odoo_tags.json del proyecto)
    cases_ok = [
        "PYL",
        "Almacenes Especialistas (PYL)",
        "Axarquía de Aislamientos. Distribución",
        "Axarquía de Aislamientos (Distribución)",
        "Axarquia de Aislamientos",
        "Axarquia de Aislamientos. Distribucion",
        "Distribuidor Axarquía",
        "Almacén de Parquet. Independiente",
        "Instalador de Parquet",
        "Almacén de Construcción. Independiente",
        "Distribuidor Oficial. Independiente",
        "Instaladores y Reformistas",
        "Empresas Constructoras",
        "Tipo de Empresa por Definir",
    ]
    cases_sin_homo = [
        "TIPO_INEXISTENTE_XYZ_123",
        "",
    ]

    n_ok = 0
    n_fail = 0
    for tipo in cases_ok:
        r = homo.homologar(tipo)
        if r.status in (HomologacionStatus.OK, HomologacionStatus.FUERA_TABLA,
                        HomologacionStatus.POR_DEFINIR):
            ok(f"homologar('{tipo[:45]}')", f"status={r.status.value}")
            n_ok += 1
        else:
            err(f"homologar('{tipo[:45]}') -> SIN_HOMOLOGACION",
                f"Tipo Odoo real no encontrado en catalogo")
            n_fail += 1

    for tipo in cases_sin_homo:
        r = homo.homologar(tipo)
        if r.status == HomologacionStatus.SIN_HOMOLOGACION:
            ok(f"homologar('{tipo[:30]}') -> SIN_HOMOLOGACION (correcto)")
        else:
            err(f"homologar('{tipo[:30]}') deberia ser SIN_HOMOLOGACION",
                f"obtenido={r.status.value}")

    print(f"  Tests homologacion: {n_ok} OK, {n_fail} fallos")

    # listar_entradas
    entries = homo.listar_entradas()
    if isinstance(entries, list) and len(entries) > 0:
        ok("listar_entradas()", f"{len(entries)} entradas")
    else:
        err("listar_entradas() vacio o invalido")


# ─────────────────────────────────────────────────────────────────────────────
# BLOQUE 10 — CommercialService integrado
# ─────────────────────────────────────────────────────────────────────────────
def bloque_10_commercial_service():
    print("\n[10] COMMERCIAL SERVICE — FLUJO INTEGRADO")
    try:
        from db.services.commercial_service import CommercialService
        cs = CommercialService(None)
        ok("CommercialService instancia sin Odoo")
        for method in ["get_pending_orders", "batch_validate_orders"]:
            if hasattr(cs, method):
                ok(f"CommercialService.{method} existe")
            else:
                err(f"CommercialService.{method} AUSENTE")
    except Exception as e:
        err("CommercialService importable", str(e))


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 70)
    print("  AUDITORIA VALIDADOR COMERCIAL BUR2000 — ENERO 2026 + Axarquia")
    print("=" * 70)

    excel_ok = bloque_1_integridad()
    df = bloque_2_lectura() if excel_ok else None
    segs = bloque_3_segmentos(df) if df is not None else set()
    if df is not None:
        bloque_4_tramos(df)
        bloque_5_dtos(df)
    bloque_6_portes()
    bloque_7_catalogo(segs)
    bloque_8_motor_validacion()
    bloque_9_homologacion()
    bloque_10_commercial_service()

    # ── Resumen final ─────────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print(f"  RESULTADO: {_ok_count} OK  |  {len(_warnings)} AVISOS  |  {len(_errors)} ERRORES")

    if _errors:
        print("\n  ERRORES CRITICOS:")
        for e in _errors:
            print(f"    [ERR] {e}")

    if _warnings:
        print("\n  AVISOS:")
        for w in _warnings:
            print(f"    [WARN] {w}")

    if not _errors and not _warnings:
        print("\n  PERFECTO: El validador esta al 100% listo para produccion.")
    elif not _errors:
        print("\n  Sin errores criticos. Revisar avisos si es necesario.")
    else:
        print("\n  HAY ERRORES — El validador NO esta listo para produccion.")

    print("=" * 70)
    return len(_errors)


if __name__ == "__main__":
    sys.exit(main())
