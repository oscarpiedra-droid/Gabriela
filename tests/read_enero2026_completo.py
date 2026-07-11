# -*- coding: utf-8 -*-
"""Extrae estructura completa: cabeceras, segmentos, todos los tramos."""
import sys, os, warnings
warnings.filterwarnings("ignore")
import openpyxl

EXCEL = os.path.join(os.path.dirname(__file__), '..', 'Nuevo', 'ENERO 2026 - Con Axarquia.xlsx')
wb = openpyxl.load_workbook(EXCEL, data_only=True)

print("HOJAS:", wb.sheetnames, "\n")

# === 1. PORTES ABRIL 2026 (completo) ===
print("="*80)
print("PORTES ABRIL 2026 (vigente desde 01/04/2026)")
print("="*80)
ws = wb['Portes Abril 2026']
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), 1):
    if any(v is not None for v in row):
        vals = [str(v).replace('\n', ' | ') if v is not None else "" for v in row]
        clean = " || ".join(v for v in vals if v.strip())
        print(f"  F{i:02d}: {clean}")

# === 2. DESCUENTOS principales — TODAS las filas ===
# La primera hoja con muchas filas es la de descuentos
print()
for sn in wb.sheetnames:
    ws2 = wb[sn]
    if ws2.max_row > 30 and sn.strip() not in ['Portes Abril 2026', 'Portes', 'XPS', 'README']:
        print("="*80)
        print(f"HOJA DESCUENTOS: '{sn}'  ({ws2.max_row} filas x {ws2.max_column} cols)")
        print("="*80)
        for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=True), 1):
            if any(v is not None for v in row):
                vals = []
                for v in row:
                    if v is None: vals.append("")
                    elif isinstance(v, float): vals.append(f"{v:.0f}" if v == int(v) else str(v))
                    else: vals.append(str(v).replace('\n', ' | ')[:80])
                print(f"  F{i:03d}: {' | '.join(vals)}")
        print()
