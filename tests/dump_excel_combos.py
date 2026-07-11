# -*- coding: utf-8 -*-
"""Imprime todas las combinaciones únicas Segmento+Familia del Excel."""
import openpyxl, os
EXCEL = os.path.join(os.path.dirname(__file__), '..', 'Nuevo', 'ENERO 2026 - Con Axarquia.xlsx')
wb = openpyxl.load_workbook(EXCEL, data_only=True)
for sn in wb.sheetnames:
    ws = wb[sn]
    if ws.max_row < 30: continue
    combos = set()
    for row in ws.iter_rows(min_row=18, max_row=ws.max_row, values_only=True):
        seg = str(row[3] or '').strip()
        fam = str(row[4] or '').strip()
        if seg and fam and '+2 GAMAS' not in fam and '+OTRA GAMA' not in fam:
            combos.add((seg, fam))
        elif seg and fam:
            combos.add((seg, f"[BONUS] {fam}"))
    if combos:
        print(f"\n=== Hoja: {sn} ===")
        for seg, fam in sorted(combos):
            print(f"  SEG={seg!r:<50} FAM={fam!r}")
