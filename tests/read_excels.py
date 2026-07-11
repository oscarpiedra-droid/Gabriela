import sys, os
sys.path.insert(0, 'app')
import openpyxl

EXCELS = {
    'PORTES': r'Descuentos\Nueva Política de Portes 2026.xlsx',
    'DESCUENTOS': r'Descuentos\Nueva tabla de descuentos ENERO 2026.xlsx',
    'ENERO_AXARQUIA': r'Nuevo\ENERO 2026 - Con Axarquia.xlsx',
}

for label, path in EXCELS.items():
    full = os.path.join(os.getcwd(), path)
    if not os.path.exists(full):
        print(f'[NOT FOUND] {label}: {path}')
        continue
    try:
        wb = openpyxl.load_workbook(full, data_only=True)
        print(f'\n{"="*70}')
        print(f'=== {label}: {os.path.basename(path)} ===')
        print(f'Hojas: {wb.sheetnames}')
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f'\n--- Hoja: {sheet_name} ({ws.max_row} filas x {ws.max_column} cols) ---')
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), values_only=True):
                if any(v is not None for v in row):
                    print(row)
    except Exception as e:
        print(f'[ERROR] {label}: {e}')
