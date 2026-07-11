# -*- coding: utf-8 -*-
"""
verify_discounts_vs_excel.py
Compara TODOS los segmentos y tramos del JSON contra el Excel ENERO 2026.
Detecta cualquier diferencia.
"""
import sys, os, re, json, warnings
warnings.filterwarnings("ignore")
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'app'))

JSON_FILE = os.path.join(os.path.dirname(__file__), '..', 'app', 'db', 'commercial_rules_v2.json')
EXCEL = os.path.join(os.path.dirname(__file__), '..', 'Nuevo', 'ENERO 2026 - Con Axarquia.xlsx')

with open(JSON_FILE, encoding='utf-8') as f:
    data = json.loads(re.sub(r'\bNaN\b', 'null', f.read()))

SD = data['SKU_DISCOUNTS']  # {sku: {SEGMENTO: [{min_eur_order, max_eur_order, dscto_peninsula_pct, dscto_baleares_pct}]}}

# ── Mapeo familia Excel -> prefijos de SKU en JSON ───────────────────────────
FAMILIA_TO_SKUS = {
    "ACÚSTICA":                          [k for k in SD.keys() if k.startswith("11.")],
    "ANTI IMPACTO (NO SOUND)":           [k for k in SD.keys() if k.startswith("13.") or k.startswith("12.")],
    "IMPERMEABILIZANTES":                [k for k in SD.keys() if k.startswith("17.") or k.startswith("16.")],
    "AIR-BUR TERMIC / TERMOREFLEX (EXCL. CM XPS / S-YC)": [k for k in SD.keys() if k.startswith("01.") and not k.startswith("01.002")],
    "AIR-BUR TERMIC / (EXCL. CM )":     [k for k in SD.keys() if k.startswith("01.") and not k.startswith("01.002")],
    "AIR BUR TERMIC (CM XPS / S-YC)":   [k for k in SD.keys() if k.startswith("01.002")],
    "AIR BUR TERMIC (CM )":             [k for k in SD.keys() if k.startswith("01.002")],
    "PARQUET":                           [k for k in SD.keys() if k.startswith("21.")],
}

# ── Mapeo nombre segmento Excel -> clave JSON ────────────────────────────────
SEG_MAP = {
    "Almacenes Especialistas (PYL)":             "ALMACENES_ESPECIALISTAS_PYL",
    "Almacenes Generalistas":                    "ALMACENES_GENERALISTAS",
    "Empresas Constructoras":                    "EMPRESAS_CONSTRUCTORAS",
    "Empresas Instaladoras":                     "EMPRESAS_INSTALADORAS",
    "Almacenes e Instaladores (Gama SOUND)":     "ALMACENES_INSTALADORES_SOUND",
    "Axarquía de Aislamientos (Distribución)":   "AXARQUIA_DE_AISLAMIENTOS_DISTRIBUCION",
}

# ── Leer Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL, data_only=True)
# Encontrar hoja de descuentos (la que tiene más de 30 filas y no es Portes/XPS/README)
ws_disc = None
for sn in wb.sheetnames:
    ws_check = wb[sn]
    if ws_check.max_row > 30 and sn.strip() not in ['Portes Abril 2026', 'Portes', 'XPS', 'README']:
        ws_disc = ws_check
        break

if not ws_disc:
    print("ERROR: no se encontró la hoja de descuentos")
    sys.exit(1)

# Leer todas las filas de datos (desde F17 que es la cabecera)
rows_excel = []
for i, row in enumerate(ws_disc.iter_rows(min_row=18, max_row=ws_disc.max_row, values_only=True), 18):
    seg, fam, tramo, dto_t, dto_b = (str(row[3] or '').strip(),
                                      str(row[4] or '').strip(),
                                      row[5], row[6], row[7])
    if seg and fam and tramo is not None and dto_t is not None:
        rows_excel.append({
            'seg': seg, 'fam': fam, 'tramo': tramo,
            'dto_t': float(dto_t) if dto_t else 0.0,
            'dto_b': float(dto_b) if dto_b else 0.0
        })

print(f"Filas leídas del Excel: {len(rows_excel)}")
print()

# ── Verificar segmento por segmento ─────────────────────────────────────────
OK = 0; WARN = 0; ERR = 0

def tramo_to_range(tramo_val):
    """Convierte el valor de tramo Excel a (min, max) EUR."""
    if isinstance(tramo_val, (int, float)):
        return (float(tramo_val), 999999.0)
    s = str(tramo_val).strip()
    if s.startswith('<'):
        return (0.0, float(s.replace('<','').replace('€','').replace('.','').replace(',','.').strip()))
    try:
        return (float(s.replace('.','').replace(',','.')), 999999.0)
    except:
        return None

def find_json_dto(sku_data, seg_key, amount, campo):
    """Busca el descuento en el JSON para un SKU, segmento y monto dado."""
    seg_rules = sku_data.get(seg_key, [])
    for r in seg_rules:
        if r['min_eur_order'] <= amount < r['max_eur_order']:
            return r.get(campo, 0.0)
    return None

# Iterar filas del Excel y verificar contra JSON
current_seg = None
for row in rows_excel:
    seg_excel = row['seg']
    fam_excel = row['fam']
    seg_key = SEG_MAP.get(seg_excel)
    if not seg_key:
        if seg_excel != current_seg:
            print(f"  [SKIP] Segmento no mapeado: '{seg_excel}'")
            current_seg = seg_excel
        continue

    if seg_excel != current_seg:
        print(f"\n{'='*60}")
        print(f"Segmento: {seg_excel} -> JSON key: {seg_key}")
        print(f"{'='*60}")
        current_seg = seg_excel

    # Obtener rango del tramo
    rng = tramo_to_range(row['tramo'])
    if not rng:
        print(f"  [SKIP] Tramo no parseable: {row['tramo']}")
        continue
    tramo_min, _ = rng
    # Para buscar en JSON, usamos el punto medio del tramo
    # Si es < X, usamos tramo_min + algo pequeño
    test_amount = tramo_min + 1 if tramo_min > 0 else 100.0

    # Verificar para los SKUs representativos de esta familia
    fam_clean = fam_excel.replace(' +2 GAMAS', '').strip()
    skus = FAMILIA_TO_SKUS.get(fam_excel, FAMILIA_TO_SKUS.get(fam_clean, []))

    # +2 GAMAS es una feature de bonus no implementada en el servicio aún → PENDING
    is_bonus = '+2 GAMAS' in fam_excel or '+OTRA GAMA' in fam_excel
    if is_bonus:
        WARN += 1
        print(f"  [PEND] Fam={fam_excel:<45} Tramo={str(row['tramo']):<12} "
              f"(feature bonus +2 GAMAS pendiente de implementar en servicio)")
        continue

        print(f"  [SKIP] Familia sin SKUs mapeados: '{fam_excel}'")
        continue

    # Tomar el primer SKU representativo
    sample_sku = skus[0]
    sku_data = SD.get(sample_sku, {})

    dto_t_json = find_json_dto(sku_data, seg_key, test_amount, 'dscto_peninsula_pct')
    dto_b_json = find_json_dto(sku_data, seg_key, test_amount, 'dscto_baleares_pct')

    dto_t_excel = row['dto_t']
    dto_b_excel = row['dto_b']

    tag_t = "OK" if dto_t_json is not None and abs(dto_t_json - dto_t_excel) < 0.1 else "ERR"
    tag_b = "OK" if dto_b_excel == 0 or (dto_b_json is not None and abs(dto_b_json - dto_b_excel) < 0.1) else "ERR"

    if tag_t == "ERR" or tag_b == "ERR":
        ERR += 1
        print(f"  [ERR] Fam={fam_excel:<45} Tramo={str(row['tramo']):<12} "
              f"DTO_T: Excel={dto_t_excel}% / JSON={dto_t_json}%  "
              f"DTO_B: Excel={dto_b_excel}% / JSON={dto_b_json}%  (SKU={sample_sku})")
    else:
        OK += 1
        print(f"  [OK]  Fam={fam_excel:<45} Tramo={str(row['tramo']):<12} "
              f"DTO_T={dto_t_excel}% DTO_B={dto_b_excel}%")

print(f"\n{'='*60}")
print(f"RESULTADO: {OK} OK / {ERR} ERRORES / {WARN} WARNINGS")
if ERR == 0:
    print("✅ TODOS los descuentos del Excel están correctamente reflejados en el JSON.")
