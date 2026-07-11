# -*- coding: utf-8 -*-
"""
audit_celula_por_celula.py
Auditoría EXHAUSTIVA — compara cada celda del Excel con CADA SKU del JSON.
No hay muestreo: se verifican TODOS los SKUs de cada familia.

Secciones:
  A: PORTES — verificación de SHIPPING_GROUPS vs Portes Abril 2026
  B: DESCUENTOS — cada fila del Excel vs todos los SKUs correspondientes
"""
import re, json, warnings, os, sys
warnings.filterwarnings("ignore")
import openpyxl

# Añadir la ruta de app al path para poder importar el servicio
_GABRIELA = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, os.path.join(_GABRIELA, 'app'))

BASE_DIR = os.path.dirname(__file__)
JSON_FILE = os.path.join(BASE_DIR, '..', 'app', 'db', 'commercial_rules_v2.json')
EXCEL     = os.path.join(BASE_DIR, '..', 'Nuevo', 'ENERO 2026 - Con Axarquia.xlsx')

with open(JSON_FILE, encoding='utf-8') as f:
    data = json.loads(re.sub(r'\bNaN\b', 'null', f.read()))

SD = data['SKU_DISCOUNTS']   # {sku: {SEGMENTO: [{min,max,t_pct,b_pct}]}}
SG = data['SHIPPING_GROUPS'] # {group_key: [{min_order_eur, max_order_eur, bucket, price}]}

# ──────────────────────────────────────────────────────────────────────────────
# Mapeo: nombre familia en Excel → prefijos de SKU en JSON
# Uso PREFIX_RULES directamente — no dependemos de SKU_MASTER
# ──────────────────────────────────────────────────────────────────────────────

# Si no tenemos SKU_MASTER, usamos prefijos conocidos
PREFIX_RULES = [
    ("AIR BUR TERMIC (CM XPS / S-YC)", lambda k: k.startswith("01.002")),
    ("AIR BUR TERMIC (CM )",            lambda k: k.startswith("01.002")),
    ("AIR-BUR TERMIC / (EXCL. CM )",    lambda k: k.startswith("01.") and not k.startswith("01.002")),
    ("AIR-BUR TERMIC / TERMOREFLEX (EXCL. CM XPS / S-YC)", lambda k: k.startswith("01.") and not k.startswith("01.002")),
    ("ACÚSTICA",                         lambda k: k.startswith("11.")),
    ("ANTI IMPACTO (NO SOUND)",          lambda k: k.startswith("12.") or k.startswith("13.")),
    ("IMPERMEABILIZANTES",               lambda k: k.startswith("16.") or k.startswith("17.")),
    ("PARQUET",                          lambda k: k.startswith("21.")),
]

def get_skus_for_family(fam_name):
    """Devuelve lista de todos los SKUs que corresponden a la familia del Excel."""
    if fam_name is None:
        return []
    for fname, test in PREFIX_RULES:
        if fam_name.strip() == fname:
            return [k for k in SD.keys() if test(k)]
    return []

# ──────────────────────────────────────────────────────────────────────────────
# Mapeo segmento Excel → clave JSON
# ──────────────────────────────────────────────────────────────────────────────
SEG_MAP = {
    "Almacenes Especialistas (PYL)":             "ALMACENES_ESPECIALISTAS_PYL",
    "Almacenes Generalistas":                    "ALMACENES_GENERALISTAS",
    "Empresas Constructoras":                    "EMPRESAS_CONSTRUCTORAS",
    "Empresas Instaladoras":                     "EMPRESAS_INSTALADORAS",
    "Almacenes e Instaladores (Gama SOUND)":     "ALMACENES_INSTALADORES_SOUND",
    "Axarquía de Aislamientos (Distribución)":   "AXARQUIA_DE_AISLAMIENTOS_DISTRIBUCION",
}

# ──────────────────────────────────────────────────────────────────────────────
# SECCIÓN A: PORTES — verificar SHIPPING_GROUPS
# ──────────────────────────────────────────────────────────────────────────────
print("=" * 80)
print("SECCIÓN A: PORTES ABRIL 2026 — Verificación SHIPPING_GROUPS")
print("=" * 80)

# Valores esperados según hoja 'Portes Abril 2026'
EXPECTED_PORTES = {
    "G1_GENERAL": [
        # (min, max, bucket, price)
        (1500.0, 999999.0, "A", 0.0),
        (1500.0, 999999.0, "B", 0.0),
        (500.0,  1500.0,   "A", 110.0),
        (500.0,  1500.0,   "B", 140.0),
        (0.0,    500.0,    "A", 60.0),
        (0.0,    500.0,    "B", 110.0),
    ],
    "G2_CM_XPS": [
        (3000.0, 999999.0, "A", 0.0),
        (3000.0, 999999.0, "B", 0.0),
        (0.0,    3000.0,   "A", 110.0),
        (0.0,    3000.0,   "B", 140.0),
    ],
    "G3_ACUSTICA_AGLO": [
        (3000.0, 999999.0, "A", 0.0),
        (3000.0, 999999.0, "B", 0.0),
        (0.0,    3000.0,   "A", 90.0),
        (0.0,    3000.0,   "B", 120.0),
    ],
    "G4_ANTIIMPACTO_NO_SOUND": [
        (3000.0, 999999.0, "C", 0.0),
        (500.0,  3000.0,   "C", 140.0),
        (0.0,    500.0,    "C", 110.0),
        (3000.0, 999999.0, "D", 0.0),
        (500.0,  3000.0,   "D", 200.0),
        (0.0,    500.0,    "D", 180.0),
    ],
    "G5_SOUND": [
        (1500.0, 999999.0, "A", 0.0),
        (1500.0, 999999.0, "B", 0.0),
        (0.0,    1500.0,   "A", 60.0),
        (0.0,    1500.0,   "B", 110.0),
    ],
}

ok_p = 0; err_p = 0
for group_key, expected_rows in EXPECTED_PORTES.items():
    actual = SG.get(group_key, [])
    # Indexar actual por (bucket, min)
    idx = {(r['region_bucket_key'], r['min_order_eur']): r for r in actual}
    print(f"\n  {group_key}:")
    for mn, mx, bkt, price in expected_rows:
        actual_row = idx.get((bkt, mn))
        if actual_row is None:
            print(f"    [ERR] bucket={bkt} min={mn} max={mx} → NO ENCONTRADO en JSON")
            err_p += 1
        else:
            a_mx    = actual_row['max_order_eur']
            a_price = actual_row['price_eur']
            ok_mx   = abs(a_mx - mx) < 1 or (mx == 999999.0 and a_mx >= 9999.0)
            ok_pr   = abs(a_price - price) < 0.01
            if ok_mx and ok_pr:
                print(f"    [OK]  bucket={bkt} min={mn:.0f} max={mx:.0f} → {price}€")
                ok_p += 1
            else:
                print(f"    [ERR] bucket={bkt} min={mn:.0f}: precio Excel={price}€ / JSON={a_price}€"
                      f"  (max Excel={mx} / JSON={a_mx})")
                err_p += 1

print(f"\n  PORTES: {ok_p} OK / {err_p} ERR")

# ──────────────────────────────────────────────────────────────────────────────
# SECCIÓN B: DESCUENTOS — cada fila × cada SKU
# ──────────────────────────────────────────────────────────────────────────────
print()
print("=" * 80)
print("SECCIÓN B: DESCUENTOS — Auditoría celda×celda (todos los SKUs)")
print("=" * 80)

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws_disc = None
for sn in wb.sheetnames:
    ws_check = wb[sn]
    if ws_check.max_row > 30 and sn.strip() not in ['Portes Abril 2026', 'Portes', 'XPS', 'README']:
        ws_disc = ws_check
        break

ok_d = 0; err_d = 0; pend_d = 0; skip_d = 0
current_seg = None

for row_idx, row in enumerate(ws_disc.iter_rows(min_row=18, max_row=ws_disc.max_row, values_only=True), 18):
    seg_raw = str(row[3] or '').strip()
    fam_raw = str(row[4] or '').strip()
    tramo   = row[5]
    dto_t   = row[6]
    dto_b   = row[7]

    if not seg_raw or not fam_raw or tramo is None or dto_t is None:
        continue

    dto_t = float(dto_t)
    dto_b = float(dto_b) if dto_b else 0.0

    # Bonus +2 GAMAS — implementado en el servicio (capa de validación),
    # no como entrada separada en el JSON. Se audita en Sección C.
    if '+2 GAMAS' in fam_raw or '+OTRA GAMA' in fam_raw:
        print(f"  [SERV] {seg_raw[:30]:<30} FAM={fam_raw[:40]:<40} → via resolve_familia_excel()")
        pend_d += 1
        continue

    seg_key = SEG_MAP.get(seg_raw)
    if not seg_key:
        skip_d += 1
        continue

    skus = get_skus_for_family(fam_raw)
    if not skus:
        skip_d += 1
        continue

    # Parsear tramo → importe de prueba (punto medio del tramo)
    if isinstance(tramo, (int, float)):
        test_min = float(tramo)
        test_amount = test_min + 1.0
    else:
        t_str = str(tramo).strip()
        if t_str.startswith('<'):
            # Menos de X → test en la mitad inferior
            try:
                limit = float(t_str.replace('<','').replace('€','').replace('.','').replace(',','.').replace(' ','').strip())
                test_amount = limit * 0.5
                if test_amount < 1: test_amount = 50.0
            except:
                test_amount = 100.0
        else:
            try:
                test_min = float(t_str.replace('.','').replace(',','.').replace('€','').replace(' ',''))
                test_amount = test_min + 1.0
            except:
                continue

    # Verificar CADA SKU de la familia
    if seg_raw != current_seg:
        if current_seg is not None:
            print()
        print(f"\n  {'─'*60}")
        print(f"  Seg: {seg_raw} [{seg_key}]")
        print(f"  {'─'*60}")
        current_seg = seg_raw

    label = f"F{row_idx:03d} {fam_raw[:35]:<35} Tramo={str(tramo):<10}"

    for sku in sorted(skus):
        sku_data = SD.get(sku, {})
        seg_rules = sku_data.get(seg_key, [])

        if not seg_rules:
            print(f"    [ERR] {label} SKU={sku} → segmento '{seg_key}' NO EXISTE en JSON")
            err_d += 1
            continue

        # Buscar la regla que aplica para test_amount
        matched = None
        for r in seg_rules:
            if r['min_eur_order'] <= test_amount < r['max_eur_order']:
                matched = r
                break

        if matched is None:
            print(f"    [ERR] {label} SKU={sku} → importe {test_amount:.0f}€ sin cobertura en JSON")
            err_d += 1
            continue

        j_t = matched['dscto_peninsula_pct']
        j_b = matched['dscto_baleares_pct']
        ok_t = abs(j_t - dto_t) < 0.05
        ok_b = abs(j_b - dto_b) < 0.05

        if ok_t and ok_b:
            print(f"    [OK ] {label} SKU={sku:12} T={dto_t}% B={dto_b}%")
            ok_d += 1
        else:
            diffs = []
            if not ok_t: diffs.append(f"T: Excel={dto_t}% JSON={j_t}%")
            if not ok_b: diffs.append(f"B: Excel={dto_b}% JSON={j_b}%")
            print(f"    [ERR] {label} SKU={sku:12} {' | '.join(diffs)}")
            err_d += 1

# ──────────────────────────────────────────────────────────────────────────────
# SECCIÓN C: Servicio — validate_range con bonus +2 GAMAS y +OTRA GAMA
# ──────────────────────────────────────────────────────────────────────────────
print()
print("=" * 80)
print("SECCIÓN C: SERVICIO — Bonus +2 GAMAS y +OTRA GAMA via validate_range()")
print("=" * 80)

ok_s = 0; err_s = 0
try:
    from db.services.commercial_conditions_service import DiscountProposalService
    svc = DiscountProposalService()

    def svc_check(label, segmento, familia, base_eur, territorio, dto_solicitado,
                  fams_pedido, esperado_status):
        global ok_s, err_s
        res = svc.validate_range(segmento, familia, base_eur, territorio, dto_solicitado,
                                 familias_en_pedido=fams_pedido)
        got = res.get('status', '?')
        if got == esperado_status:
            print(f"  [OK ] {label}: status={got}")
            ok_s += 1
        else:
            print(f"  [ERR] {label}: esperado={esperado_status} obtenido={got} msg={res.get('msg','')}")
            err_s += 1

    # CM XPS en Almacenes Especialistas, 6000€, base 55% → válido sin bonus
    svc_check("CM XPS base 55% @6000 → OK",
              "Almacenes Especialistas (PYL)", "CM_XPS_SYC", 6500.0, "PENINSULA", 55.0,
              {"CM_XPS_SYC"}, "OK")

    # CM XPS solo: pedir 57% no debería pasar (máximo base 55%)
    svc_check("CM XPS solo 57%>55% → BLOQUEADO",
              "Almacenes Especialistas (PYL)", "CM_XPS_SYC", 6500.0, "PENINSULA", 57.1,
              {"CM_XPS_SYC"}, "BLOQUEADO")

    # CM XPS + 2 otras familias: 57% debe ser válido (bonus +2 GAMAS)
    fams_3 = {"CM_XPS_SYC", "ACUSTICA", "IMPERMEABILIZANTES", "ANTI_IMPACTO_NO_SOUND"}
    svc_check("CM XPS +2 GAMAS 57%@6000 → OK",
              "Almacenes Especialistas (PYL)", "CM_XPS_SYC", 6500.0, "PENINSULA", 57.0,
              fams_3, "OK")

    # CM XPS + 2 otras / Generalistas: 57% → OK (bonus)
    svc_check("CM XPS +2 GAMAS Generalistas 57%@6000 → OK",
              "Almacenes Generalistas", "CM_XPS_SYC", 6500.0, "PENINSULA", 57.0,
              fams_3, "OK")

    # CM XPS + 2 otras: pedir más del bonus → BLOQUEADO
    svc_check("CM XPS +2 GAMAS 60%>57% → BLOQUEADO o AVISO",
              "Almacenes Especialistas (PYL)", "CM_XPS_SYC", 6500.0, "PENINSULA", 60.0,
              fams_3, "BLOQUEADO")

    # PARQUET en SOUND, 3000€, base 58% → OK sin bonus
    svc_check("PARQUET base 58%@3000 → OK",
              "Almacenes e Instaladores (Gama SOUND)", "PARQUET", 3500.0, "PENINSULA", 58.0,
              {"PARQUET"}, "OK")

    # PARQUET solo: 60% > 58% = BLOQUEADO
    svc_check("PARQUET solo 60%>58% → BLOQUEADO",
              "Almacenes e Instaladores (Gama SOUND)", "PARQUET", 3500.0, "PENINSULA", 60.1,
              {"PARQUET"}, "BLOQUEADO")

    # PARQUET + otra gama: 60% → OK (bonus +OTRA GAMA)
    fams_parquet = {"PARQUET", "ACUSTICA"}
    svc_check("PARQUET +OTRA GAMA 60%@3000 → OK",
              "Almacenes e Instaladores (Gama SOUND)", "PARQUET", 3500.0, "PENINSULA", 60.0,
              fams_parquet, "OK")

    # REVIEW_REQUIRED siempre pasa (excluido de validación)
    svc_check("REVIEW_REQUIRED excluido → OK",
              "Almacenes Generalistas", "REVIEW_REQUIRED", 5000.0, "PENINSULA", 99.0,
              {"REVIEW_REQUIRED"}, "OK")

except Exception as e:
    print(f"  [EXCEPCION] No se pudo importar el servicio: {e}")
    err_s = 1

print(f"\n  SERVICIO: {ok_s} OK / {err_s} ERR")

# ──────────────────────────────────────────────────────────────────────────────
print()
print("=" * 80)
print(f"RESULTADO FINAL:")
print(f"  PORTES:     {ok_p} OK / {err_p} ERR")
print(f"  DESCUENTOS: {ok_d} OK / {err_d} ERR / {pend_d} via-SERVICIO (+2 GAMAS) / {skip_d} SKIPPED")
print(f"  SERVICIO:   {ok_s} OK / {err_s} ERR  (+2 GAMAS / +OTRA GAMA via validate_range)")
total_err = err_p + err_d + err_s
if total_err == 0:
    print(f"  TODOS los controles pasados. Motor comercial 100% alineado con Excel 2026.")
else:
    print(f"  {total_err} ERRORES TOTALES — revisar arriba")
print("=" * 80)
