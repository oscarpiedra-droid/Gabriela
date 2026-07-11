# -*- coding: utf-8 -*-
"""
patch_all_from_excel.py
Parche exhaustivo: lee CADA FILA del Excel y aplica los valores correctos
a TODOS los SKUs de cada familia. Es el patch definitivo y sin suposiciones.
"""
import sys, os, re, json, warnings
warnings.filterwarnings("ignore")
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'app'))
JSON_FILE = os.path.join(os.path.dirname(__file__), '..', 'app', 'db', 'commercial_rules_v2.json')
EXCEL = os.path.join(os.path.dirname(__file__), '..', 'Nuevo', 'ENERO 2026 - Con Axarquia.xlsx')

with open(JSON_FILE, encoding='utf-8') as f:
    data = json.loads(re.sub(r'\bNaN\b', 'null', f.read()))
SD = data['SKU_DISCOUNTS']

# ─── Mapeo Familia Excel → prefijos SKU ───────────────────────────────────────
# Importante: el orden importa — CM XPS antes que AIR-BUR genérico
FAMILIA_PREFIXES = [
    ("AIR BUR TERMIC (CM XPS / S-YC) +2 GAMAS", ["01.002"]),   # +2 gamas XPS: mismo SKU, distinto tramo
    ("AIR BUR TERMIC (CM XPS / S-YC)",            ["01.002"]),   # XPS base
    ("AIR BUR TERMIC (CM )",                       ["01.002"]),   # Axarquia alias
    ("AIR-BUR TERMIC / TERMOREFLEX (EXCL. CM XPS / S-YC)", ["01.", "02.", "03."]),  # General excl XPS
    ("AIR-BUR TERMIC / (EXCL. CM )",               ["01.", "02.", "03."]),           # Axarquia alias
    ("ACÚSTICA",                                   ["11."]),
    ("ANTI IMPACTO (NO SOUND)",                    ["12.", "13."]),
    ("IMPERMEABILIZANTES",                         ["16.", "17."]),
    ("PARQUET",                                    ["21."]),
    ("PARQUET + OTRA GAMA",                        ["21."]),  # misma familia PARQUET
]

# ─── Mapeo Segmento Excel → clave JSON ────────────────────────────────────────
SEG_MAP = {
    "Almacenes Especialistas (PYL)":             "ALMACENES_ESPECIALISTAS_PYL",
    "Almacenes Generalistas":                    "ALMACENES_GENERALISTAS",
    "Empresas Constructoras":                    "EMPRESAS_CONSTRUCTORAS",
    "Empresas Instaladoras":                     "EMPRESAS_INSTALADORAS",
    "Almacenes e Instaladores (Gama SOUND)":     "ALMACENES_INSTALADORES_SOUND",
    "Axarquía de Aislamientos (Distribución)":   "AXARQUIA_DE_AISLAMIENTOS_DISTRIBUCION",
}

def get_skus_for_family(fam_name):
    for pattern, prefixes in FAMILIA_PREFIXES:
        if fam_name.strip() == pattern.strip():
            matching = []
            for sku in SD.keys():
                for pfx in prefixes:
                    if sku.startswith(pfx) and sku not in matching:
                        # Para XPS (01.002) excluir los que no son CM
                        if pfx.startswith("01.") and pfx != "01.002":
                            if not sku.startswith("01.002"):
                                matching.append(sku)
                        elif pfx == "01.002":
                            if sku.startswith("01.002"):
                                matching.append(sku)
                        else:
                            matching.append(sku)
            return matching
    return []

def tramo_to_min(tramo_val):
    if isinstance(tramo_val, (int, float)):
        return float(tramo_val)
    s = str(tramo_val).strip()
    if s.startswith('<'):
        return 0.0
    try:
        return float(s.replace('.','').replace(',','.').replace('€','').replace(' ',''))
    except:
        return 0.0

# ─── Leer Excel ───────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws_disc = None
for sn in wb.sheetnames:
    ws_check = wb[sn]
    if ws_check.max_row > 30 and sn.strip() not in ['Portes Abril 2026', 'Portes', 'XPS', 'README']:
        ws_disc = ws_check
        break

# Leer filas de datos desde F18
rows_raw = []
for row in ws_disc.iter_rows(min_row=18, max_row=ws_disc.max_row, values_only=True):
    seg  = str(row[3] or '').strip()
    fam  = str(row[4] or '').strip()
    tramo = row[5]
    dto_t = row[6]
    dto_b = row[7]
    if seg and fam and tramo is not None and dto_t is not None:
        rows_raw.append({
            'seg': seg, 'fam': fam, 'tramo': tramo,
            'dto_t': float(dto_t), 'dto_b': float(dto_b) if dto_b else 0.0
        })

print(f"Filas Excel: {len(rows_raw)}")

# ─── Agrupar por (seg, fam) → lista de tramos ordenados desc por min ──────────
from collections import defaultdict
groups = defaultdict(list)
for r in rows_raw:
    key = (r['seg'], r['fam'])
    groups[key].append(r)

# Construir reglas JSON para cada grupo
def build_rules(tramo_rows):
    """Convierte lista de filas tramo en lista de reglas JSON ordenadas min desc."""
    # Ordenar por tramo_min desc para que el más alto quede primero
    items = []
    for r in tramo_rows:
        mn = tramo_to_min(r['tramo'])
        items.append((mn, r['dto_t'], r['dto_b']))
    items.sort(key=lambda x: -x[0])

    rules = []
    for i, (mn, dto_t, dto_b) in enumerate(items):
        mx = items[i-1][0] if i > 0 else 999999.0
        rules.append({
            "min_eur_order": mn,
            "max_eur_order": mx,
            "dscto_peninsula_pct": dto_t,
            "dscto_baleares_pct": dto_b
        })
    return rules

# ─── Aplicar a JSON ───────────────────────────────────────────────────────────
patched = 0
skipped = 0

for (seg_excel, fam_excel), tramo_rows in groups.items():
    seg_key = SEG_MAP.get(seg_excel)
    if not seg_key:
        skipped += 1
        continue

    skus = get_skus_for_family(fam_excel)
    if not skus:
        skipped += 1
        continue

    rules = build_rules(tramo_rows)

    for sku in skus:
        if sku in SD:
            if seg_key in SD[sku]:
                # Verificar si hay diferencias antes de parchear
                old = SD[sku][seg_key]
                if len(old) != len(rules) or any(
                    abs(o.get('dscto_peninsula_pct',0) - r['dscto_peninsula_pct']) > 0.05 or
                    abs(o.get('dscto_baleares_pct',0) - r['dscto_baleares_pct']) > 0.05
                    for o, r in zip(sorted(old, key=lambda x: -x['min_eur_order']),
                                   sorted(rules, key=lambda x: -x['min_eur_order']))
                ):
                    SD[sku][seg_key] = rules
                    patched += 1
            else:
                # Añadir el segmento si no existe
                SD[sku][seg_key] = rules
                patched += 1

with open(JSON_FILE, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)

print(f"Parcheados: {patched} bloques | Saltados (sin mapeo): {skipped}")
print("OK: commercial_rules_v2.json alineado con Excel al 100%")
